"""
Excel consolidado: MEJOR VALOR POR INSTANCIA Y METAHEURÍSTICA, con parámetros.

Una hoja por metaheurística (SA, TS, RTS, ABC, CS, VDO) con las 87 instancias
(23 pequeñas gdb/kshs + 64 val/gdb/egl) que tuvieron corrida, más una hoja
"Resumen" con el mejor global por instancia (ganador entre las 6) y una hoja
"Comparativa" con gap medio / BKS alcanzadas / victorias por MH.

Reutiliza las MISMAS funciones de selección que los generadores de tablas
LaTeX (_gen_tablas_mh_20260713.py, _gen_tabla_cuckoo_20260712.py) para que
el Excel sea consistente con lo ya publicado: la corrida de menor costo por
(MH, instancia), desempate por menor tiempo.

Salida: resultados/MetaCARP_mejores_resultados_20260715.xlsx

Uso:
    python scripts/_gen_excel_resumen_20260715.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ / "scripts"))

from run_val_egl_20260710 import INSTANCIAS_VAL_GDB, INSTANCIAS_EGL  # noqa: E402
from _gen_tabla_cuckoo_20260712 import (  # noqa: E402
    INSTANCIAS_SMALL, LETRA_APPROACH, _deltas, _mejor_por_instancia,
)
from _gen_tablas_mh_20260713 import MHS, _mejores as _mejores_generico  # noqa: E402

DESTINO = RAIZ / "resultados" / "MetaCARP_mejores_resultados_20260715.xlsx"
ORDEN_INSTANCIAS = list(INSTANCIAS_SMALL) + list(INSTANCIAS_VAL_GDB) + list(INSTANCIAS_EGL)

# ------------------------------------------------------------------
# Estilos
# ------------------------------------------------------------------
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FILL_BKS = PatternFill("solid", fgColor="C6EFCE")
FONT_BKS = Font(color="006100", bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_width=8, max_width=60):
    for col_cells in ws.columns:
        letra = get_column_letter(col_cells[0].column)
        largo = max((len(str(c.value)) for c in col_cells if c.value is not None),
                    default=0)
        ws.column_dimensions[letra].width = max(min_width, min(max_width, largo + 2))


def _escribir_encabezado(ws, columnas: list[str]) -> None:
    ws.append(columnas)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    ws.freeze_panes = "A2"


# ------------------------------------------------------------------
# Recolección por MH (reutiliza los selectores ya validados)
# ------------------------------------------------------------------

def _fila_sa_ts_rts_abc(clave_mh: str, deltas: dict[str, float]) -> list[dict]:
    """SA/TS/RTS/ABC: usa MHS[clave_mh] + _mejores_generico (junio+julio)."""
    spec = MHS[clave_mh]
    mejores = _mejores_generico(spec)
    filas = []
    for inst in ORDEN_INSTANCIAS:
        b = mejores.get(inst)
        if b is None:
            continue
        delta = deltas.get(inst, 0.0)
        bks = float(b["bks_referencia"])
        costo = b["_costo"] + delta
        gap = (costo - bks) / bks * 100.0
        fila = {
            "instancia": inst,
            "bks": bks,
            "costo": round(costo, 2),
            "gap_%": round(gap, 3),
            "tiempo_s": round(b["_tiempo"], 2),
            "repeticion": b.get("repeticion", ""),
            "semilla": b.get("semilla", ""),
            "origen": b["_origen"],
        }
        for etiqueta, col, fmt in spec["params"]:
            # etiqueta viene en LaTeX ($T_0$, etc.); usamos el nombre de
            # columna CSV como encabezado Excel, más legible sin editar.
            valor = b.get(col)
            fila[col] = fmt(valor) if valor not in (None, "") else None
        fila["p_inter"] = b.get("p_inter")
        filas.append(fila)
    return filas


def _fila_cs(deltas: dict[str, float]) -> list[dict]:
    small = _mejor_por_instancia(
        [("experimentos_costo_fixed/cuckoo_*/final/*.csv", "?")])
    grandes = _mejor_por_instancia(
        [("experimentos_val_egl_20260710/cs_campana/cs_minigrid/"
          "corrida_*/final/*.csv", "W")])
    mejores = {**small, **grandes}
    filas = []
    for inst in ORDEN_INSTANCIAS:
        b = mejores.get(inst)
        if b is None:
            continue
        delta = deltas.get(inst, 0.0)
        bks = b["bks"]
        costo = b["costo"] + delta
        gap = (costo - bks) / bks * 100.0
        filas.append({
            "instancia": inst,
            "bks": bks,
            "costo": round(costo, 2),
            "gap_%": round(gap, 3),
            "tiempo_s": round(b["tiempo"], 2),
            "num_nidos": b["nidos"],
            "pa_abandono": b["pa"],
            "lambda_levy": b["beta"],
            "factor_pasos": b["fpasos"],
            "p_inter": b["p_inter"],
            "origen": b["origen"],
        })
    return filas


def _fila_vdo(deltas: dict[str, float]) -> list[dict]:
    spec = MHS["vdo"]
    mejores = _mejores_generico(spec)
    filas = []
    for inst in ORDEN_INSTANCIAS:
        b = mejores.get(inst)
        if b is None:
            continue
        delta = deltas.get(inst, 0.0)
        bks = float(b["bks_referencia"])
        costo = b["_costo"] + delta
        gap = (costo - bks) / bks * 100.0
        filas.append({
            "instancia": inst,
            "bks": bks,
            "costo": round(costo, 2),
            "gap_%": round(gap, 3),
            "tiempo_s": round(b["_tiempo"], 2),
            "amplitud_inicial_A0": (
                round(float(b["amplitud_inicial_efectiva"]), 3)
                if b.get("amplitud_inicial_efectiva") not in (None, "") else None),
            "sigma": (round(float(b["sigma_efectivo"]), 3)
                      if b.get("sigma_efectivo") not in (None, "") else None),
            "gamma": b.get("gamma"),
            "L_por_nivel": (round(float(b["iteraciones_por_nivel_L"]))
                            if b.get("iteraciones_por_nivel_L") not in (None, "") else None),
            "p_inter": b.get("p_inter"),
            "origen": b["_origen"],
        })
    return filas


# ------------------------------------------------------------------
# Escritura de hojas
# ------------------------------------------------------------------

def _hoja_mh(wb: Workbook, nombre: str, filas: list[dict]) -> None:
    ws = wb.create_sheet(nombre)
    if not filas:
        ws.append(["(sin datos)"])
        return
    columnas = list(filas[0].keys())
    _escribir_encabezado(ws, columnas)
    idx_costo = columnas.index("costo") + 1
    idx_bks = columnas.index("bks") + 1
    for fila in filas:
        ws.append([fila.get(c) for c in columnas])
        r = ws.max_row
        if fila["costo"] <= fila["bks"] + 1e-6:
            for c in (idx_costo, idx_bks):
                ws.cell(row=r, column=c).fill = FILL_BKS
                ws.cell(row=r, column=c).font = FONT_BKS
    _autofit(ws)


def _hoja_resumen(wb: Workbook, por_mh: dict[str, list[dict]]) -> None:
    ws = wb.create_sheet("Resumen (mejor global)")
    columnas = ["instancia", "bks", "mejor_costo", "gap_%", "metaheuristica"]
    _escribir_encabezado(ws, columnas)
    indices = {mh: {f["instancia"]: f for f in filas} for mh, filas in por_mh.items()}
    for inst in ORDEN_INSTANCIAS:
        candidatos = [(mh, d[inst]) for mh, d in indices.items() if inst in d]
        if not candidatos:
            continue
        mh_gan, fila_gan = min(candidatos, key=lambda kv: kv[1]["costo"])
        ws.append([inst, fila_gan["bks"], fila_gan["costo"], fila_gan["gap_%"], mh_gan])
        r = ws.max_row
        if fila_gan["costo"] <= fila_gan["bks"] + 1e-6:
            for c in (2, 3):
                ws.cell(row=r, column=c).fill = FILL_BKS
                ws.cell(row=r, column=c).font = FONT_BKS
    _autofit(ws)


def _hoja_comparativa(wb: Workbook, por_mh: dict[str, list[dict]]) -> None:
    import statistics
    ws = wb.create_sheet("Comparativa por MH", 0)  # primera hoja
    columnas = ["metaheuristica", "n_instancias", "gap_medio_%",
                "gap_mediana_%", "BKS_alcanzadas", "victorias_(mejor_o_empate)"]
    _escribir_encabezado(ws, columnas)
    indices = {mh: {f["instancia"]: f for f in filas} for mh, filas in por_mh.items()}
    for mh, filas in por_mh.items():
        if not filas:
            continue
        gaps = [f["gap_%"] for f in filas]
        bks = sum(1 for f in filas if f["costo"] <= f["bks"] + 1e-6)
        victorias = 0
        for inst in ORDEN_INSTANCIAS:
            candidatos = [(m, d[inst]) for m, d in indices.items() if inst in d]
            if not candidatos:
                continue
            mh_gan, _ = min(candidatos, key=lambda kv: kv[1]["costo"])
            if mh_gan == mh:
                victorias += 1
        ws.append([mh, len(filas), round(statistics.mean(gaps), 3),
                   round(statistics.median(gaps), 3), f"{bks}/{len(filas)}",
                   victorias])
    _autofit(ws)


def main() -> None:
    deltas = _deltas()
    por_mh = {
        "SA": _fila_sa_ts_rts_abc("sa", deltas),
        "TS": _fila_sa_ts_rts_abc("ts", deltas),
        "RTS": _fila_sa_ts_rts_abc("rts", deltas),
        "ABC": _fila_sa_ts_rts_abc("abc", deltas),
        "CS": _fila_cs(deltas),
        "VDO": _fila_vdo(deltas),
    }

    wb = Workbook()
    wb.remove(wb.active)  # quita la hoja default "Sheet"
    _hoja_comparativa(wb, por_mh)
    _hoja_resumen(wb, por_mh)
    for mh, filas in por_mh.items():
        _hoja_mh(wb, mh, filas)
        print(f"[{mh}] {len(filas)} instancias")

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINO)
    print(f"\nGuardado: {DESTINO}")


if __name__ == "__main__":
    main()
