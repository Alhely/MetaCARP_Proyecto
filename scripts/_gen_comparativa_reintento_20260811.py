"""
Comparativa ANTES/DESPUÉS del re-intento dirigido (2026-08-05 -> 2026-08-11).

Para cada (MH, instancia) de resultados/instancias_gap_mayor_1pct_20260805.csv,
busca el mejor resultado nuevo en experimentos_reintento_20260805/ y lo
compara contra el mejor original (Excel de julio). Reporta el mejor de los
dos (nunca empeora: si el re-intento no mejoró, se conserva el original) y
cuenta cuántas instancias cruzaron el umbral de 1%.

Salidas:
  resultados/comparativa_reintento_20260811.csv   (detalle por fila)
  resultados/MetaCARP_mejores_resultados_20260811.xlsx (Excel actualizado,
      misma estructura del 20260715 + hoja "Antes vs Despues del reintento")

Uso:
    python scripts/_gen_comparativa_reintento_20260811.py
"""
from __future__ import annotations

import csv
import glob
import statistics
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ / "scripts"))

from _gen_tabla_cuckoo_20260712 import _deltas  # noqa: E402

LISTA_PROBLEMA = RAIZ / "resultados" / "instancias_gap_mayor_1pct_20260805.csv"
DIR_REINTENTO = RAIZ / "experimentos_reintento_20260805"
XLSX_ORIGINAL = RAIZ / "resultados" / "MetaCARP_mejores_resultados_20260715.xlsx"
XLSX_SALIDA = RAIZ / "resultados" / "MetaCARP_mejores_resultados_20260811.xlsx"
CSV_SALIDA = RAIZ / "resultados" / "comparativa_reintento_20260811.csv"

ETIQUETA_A_CLAVE = {"SA": "sa", "TS": "ts", "RTS": "rts", "ABC": "abc",
                    "VDO": "vdo", "CS": "cs"}


def _mejor_por_instancia_dirigido(mh: str, deltas: dict) -> dict[str, dict]:
    """Mejor fila nueva por instancia para sa/ts/rts/abc/vdo."""
    mejores: dict[str, dict] = {}
    for ruta in glob.glob(str(DIR_REINTENTO / mh / "final" / "*.csv")):
        with open(ruta, newline="") as fh:
            for fila in csv.DictReader(fh):
                inst = fila["instancia"]
                bks = float(fila["bks_referencia"])
                costo = float(fila["mejor_costo"]) + deltas.get(inst, 0.0)
                a = mejores.get(inst)
                if a is None or costo < a["costo"]:
                    mejores[inst] = {"costo": costo, "bks": bks,
                                     "p_inter": fila.get("p_inter", "")}
    return mejores


def _mejor_por_instancia_cs(deltas: dict) -> dict[str, dict]:
    mejores: dict[str, dict] = {}
    for ruta in glob.glob(str(DIR_REINTENTO / "cs_minigrid" / "corrida_*" / "final" / "*.csv")):
        with open(ruta, newline="") as fh:
            for fila in csv.DictReader(fh):
                inst = fila["instancia"]
                bks = float(fila["bks_referencia"])
                costo = float(fila["mejor_costo"]) + deltas.get(inst, 0.0)
                a = mejores.get(inst)
                if a is None or costo < a["costo"]:
                    mejores[inst] = {"costo": costo, "bks": bks,
                                     "factor_pasos": fila.get("factor_pasos", "")}
    return mejores


def main() -> None:
    deltas = _deltas()
    nuevos = {mh: _mejor_por_instancia_dirigido(mh, deltas)
             for mh in ("sa", "ts", "rts", "abc", "vdo")}
    nuevos["cs"] = _mejor_por_instancia_cs(deltas)

    filas_salida = []
    with open(LISTA_PROBLEMA, newline="") as fh:
        for fila in csv.DictReader(fh):
            mh = ETIQUETA_A_CLAVE[fila["metaheuristica"]]
            inst = fila["instancia"]
            gap_antes = float(fila["gap_%"])
            costo_antes = float(fila["costo_actual"])
            bks = float(fila["bks"])

            nuevo = nuevos[mh].get(inst)
            if nuevo is not None and nuevo["costo"] < costo_antes:
                costo_despues = nuevo["costo"]
                mejoro = True
            else:
                costo_despues = costo_antes
                mejoro = False
            gap_despues = (costo_despues - bks) / bks * 100.0

            filas_salida.append({
                "metaheuristica": fila["metaheuristica"], "instancia": inst,
                "familia": fila["familia"], "tier": fila["tier"], "bks": bks,
                "costo_antes": round(costo_antes, 2), "gap_antes_%": round(gap_antes, 3),
                "costo_nuevo_probado": round(nuevo["costo"], 2) if nuevo else None,
                "costo_despues": round(costo_despues, 2), "gap_despues_%": round(gap_despues, 3),
                "mejoro": mejoro,
                "cruzo_a_<=1%": gap_despues <= 1.0,
            })

    with open(CSV_SALIDA, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(filas_salida[0].keys()))
        w.writeheader()
        w.writerows(filas_salida)
    print(f"Detalle escrito: {CSV_SALIDA}")

    # --- Resumen en consola por MH ---
    mhs = ["SA", "TS", "RTS", "ABC", "CS", "VDO"]
    print(f"\n{'MH':6}{'n':>5}{'mejoraron':>11}{'cruzaron<=1%':>14}"
          f"{'gap_medio_antes':>17}{'gap_medio_despues':>19}")
    resumen_mh = {}
    for mh in mhs:
        sub = [f for f in filas_salida if f["metaheuristica"] == mh]
        if not sub:
            continue
        n_mejoro = sum(1 for f in sub if f["mejoro"])
        n_cruzo = sum(1 for f in sub if f["cruzo_a_<=1%"])
        gm_antes = statistics.mean(f["gap_antes_%"] for f in sub)
        gm_despues = statistics.mean(f["gap_despues_%"] for f in sub)
        resumen_mh[mh] = (len(sub), n_mejoro, n_cruzo, gm_antes, gm_despues)
        print(f"{mh:6}{len(sub):5}{n_mejoro:11}{n_cruzo:14}"
              f"{gm_antes:17.2f}{gm_despues:19.2f}")

    tot_n = sum(v[0] for v in resumen_mh.values())
    tot_mejoro = sum(v[1] for v in resumen_mh.values())
    tot_cruzo = sum(v[2] for v in resumen_mh.values())
    print(f"\nTOTAL: {tot_n} filas problema | {tot_mejoro} mejoraron "
          f"({100*tot_mejoro/tot_n:.1f}%) | {tot_cruzo} cruzaron a <=1% "
          f"({100*tot_cruzo/tot_n:.1f}%)")

    # ------------------------------------------------------------------
    # Excel actualizado: parte del original, reemplaza costo/gap donde
    # hubo mejora, agrega hoja "Antes vs Despues del reintento".
    # ------------------------------------------------------------------
    wb = load_workbook(XLSX_ORIGINAL)
    idx_por_mh_inst = {}
    for f in filas_salida:
        idx_por_mh_inst[(f["metaheuristica"], f["instancia"])] = f

    for mh in ("SA", "TS", "RTS", "ABC", "CS", "VDO"):
        if mh not in wb.sheetnames:
            continue
        ws = wb[mh]
        cols = [c.value for c in ws[1]]
        i_inst = cols.index("instancia")
        i_costo = cols.index("costo")
        i_gap = cols.index("gap_%")
        i_bks = cols.index("bks")
        for row in ws.iter_rows(min_row=2):
            inst = row[i_inst].value
            f = idx_por_mh_inst.get((mh, inst))
            if f is None or not f["mejoro"]:
                continue
            row[i_costo].value = f["costo_despues"]
            row[i_gap].value = f["gap_despues_%"]
            bks = row[i_bks].value
            if f["costo_despues"] <= bks + 1e-6:
                row[i_costo].fill = PatternFill("solid", fgColor="C6EFCE")
                row[i_costo].font = Font(color="006100", bold=True)

    # Hoja nueva de comparativa.
    ws_cmp = wb.create_sheet("Antes vs Despues del reintento")
    campos = list(filas_salida[0].keys())
    ws_cmp.append(campos)
    for cell in ws_cmp[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws_cmp.freeze_panes = "A2"
    i_cruzo = campos.index("cruzo_a_<=1%")
    i_mejoro = campos.index("mejoro")
    for f in filas_salida:
        ws_cmp.append([f[c] for c in campos])
        r = ws_cmp.max_row
        if f["cruzo_a_<=1%"]:
            for c in range(1, len(campos) + 1):
                ws_cmp.cell(row=r, column=c).fill = PatternFill("solid", fgColor="C6EFCE")
        elif f["mejoro"]:
            for c in range(1, len(campos) + 1):
                ws_cmp.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFEB9C")
    for col_cells in ws_cmp.columns:
        letra = get_column_letter(col_cells[0].column)
        largo = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws_cmp.column_dimensions[letra].width = max(8, min(30, largo + 2))

    # Hoja de resumen del reintento (primera posición).
    ws_res = wb.create_sheet("Resumen reintento", 0)
    ws_res.append(["metaheuristica", "n_problema", "mejoraron", "cruzaron_a_<=1%",
                   "gap_medio_antes_%", "gap_medio_despues_%"])
    for cell in ws_res[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for mh, (n, nm, nc, ga, gd) in resumen_mh.items():
        ws_res.append([mh, n, nm, nc, round(ga, 3), round(gd, 3)])
    ws_res.append(["TOTAL", tot_n, tot_mejoro, tot_cruzo, "", ""])
    for col_cells in ws_res.columns:
        letra = get_column_letter(col_cells[0].column)
        largo = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws_res.column_dimensions[letra].width = max(10, min(30, largo + 2))

    wb.save(XLSX_SALIDA)
    print(f"\nExcel actualizado: {XLSX_SALIDA}")


if __name__ == "__main__":
    main()
